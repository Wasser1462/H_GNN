# Author: zyw
# Date: 2024-12-20
# Description: check if the adjacency matrix is valid
import os
import openpyxl
import numpy as np
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

folder_path = '/data1/zengyongwang/test/H_GNN/data/train'
def delete_first_row_and_col(ws, delete_row=True, delete_col=True):
    if delete_row:
        ws.delete_rows(1)
    if delete_col:
        ws.delete_cols(1)

def is_adjacency_matrix(matrix):
    np_matrix = np.array(matrix)
    if np_matrix.shape[0] != np_matrix.shape[1]:
        return False
    upper_triangle = np_matrix[np.triu_indices_from(np_matrix, k=1)]
    lower_triangle = np_matrix[np.tril_indices_from(np_matrix, k=-1)]
    return np.all(upper_triangle == lower_triangle)

for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        delete_first_row_and_col(ws, delete_row=True, delete_col=True)
        modified_file_path = os.path.join(folder_path, f'modified_{filename}')
        wb.save(modified_file_path)

        matrix = []
        for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
            matrix.append([cell.value for cell in row])

        if ws.max_row != ws.max_column:
            logging.info(f'{filename} is not a square matrix.')
        # else:
        #     if is_adjacency_matrix(matrix):
        #         logging.info(f'{filename} is a valid adjacency matrix.')
        #     else:
        #         logging.info(f'{filename} is not a valid adjacency matrix.')

        logging.info(f'{filename} has been processed and saved as {modified_file_path}')
